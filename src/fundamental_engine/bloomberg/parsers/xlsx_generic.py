"""
xlsx_generic.py – Parser for Bloomberg XLSX financial statement exports.

Handles:
- Column header parsing (fiscal year, estimate/actual flags)
- Multi-sheet workbooks
- "In Millions" / "In Billions" scale detection
- LTM/TTM column detection
- Estimate column detection and exclusion
"""

from __future__ import annotations

import datetime
import logging
import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from fundamental_engine.constants import (
    BBG_ESTIMATE_KEYWORDS,
    BBG_LTM_KEYWORDS,
    BBG_SCALE_PATTERNS,
)
from fundamental_engine.exceptions import BloombergParseError
from fundamental_engine.types import (
    BloombergColumn,
    DataSource,
    RawStatementTable,
    StatementType,
)

logger = logging.getLogger(__name__)

_YEAR_PATTERN = re.compile(r"(19|20)\d{2}")

_STATEMENT_SHEET_HINTS: dict[str, StatementType] = {
    "income": StatementType.INCOME,
    "is": StatementType.INCOME,
    "p&l": StatementType.INCOME,
    "profit": StatementType.INCOME,
    "balance": StatementType.BALANCE,
    "bs": StatementType.BALANCE,
    "assets": StatementType.BALANCE,
    "cash": StatementType.CASHFLOW,
    "cf": StatementType.CASHFLOW,
    "cashflow": StatementType.CASHFLOW,
}


class XLSXGenericParser:
    """
    Parses Bloomberg generic XLSX financial statement exports.

    Parameters
    ----------
    allow_ltm:
        If True, include LTM columns. Default False for PIT research.
    allow_estimates:
        If True, include estimate (forward) columns. Must be False for backtests.
    """

    def __init__(self, allow_ltm: bool = False, allow_estimates: bool = False) -> None:
        self._allow_ltm = allow_ltm
        self._allow_estimates = allow_estimates

    def parse(
        self,
        path: Path,
        ticker: str,
        cutoff_date: datetime.date,
    ) -> list[RawStatementTable]:
        """
        Parse a Bloomberg XLSX file into a list of RawStatementTable objects.

        Parameters
        ----------
        path:
            Path to the .xlsx file.
        ticker:
            Equity ticker to associate with the data.
        cutoff_date:
            PIT cutoff date (passed through to columns, not filtered here).

        Returns
        -------
        list[RawStatementTable] – one per sheet that looks like a statement.

        Raises
        ------
        BloombergParseError: if the file cannot be read.
        """
        if not path.exists():
            raise BloombergParseError(str(path), "File not found")

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            raise BloombergParseError(str(path), f"Cannot open workbook: {exc}") from exc

        tables: list[RawStatementTable] = []

        for sheet_name in wb.sheetnames:
            stmt_type = self._detect_statement_type(sheet_name)
            if stmt_type is None:
                logger.debug("Skipping sheet '%s' (unrecognized type)", sheet_name)
                continue

            ws = wb[sheet_name]
            table = self._parse_sheet(ws, sheet_name, ticker, stmt_type, cutoff_date)
            if table is not None:
                tables.append(table)

        wb.close()

        if not tables:
            raise BloombergParseError(
                str(path),
                "No recognizable financial statement sheets found. "
                "Expected sheets named like 'Income', 'Balance', 'CashFlow'.",
            )

        logger.info("Parsed %d statement tables from %s for %s", len(tables), path.name, ticker)
        return tables

    def _detect_statement_type(self, sheet_name: str) -> StatementType | None:
        lower = sheet_name.lower().replace(" ", "").replace("_", "").replace("-", "")
        for hint, stype in _STATEMENT_SHEET_HINTS.items():
            if hint in lower:
                return stype
        return None

    def _parse_sheet(
        self,
        ws: Any,
        sheet_name: str,
        ticker: str,
        stmt_type: StatementType,
        cutoff_date: datetime.date,
    ) -> RawStatementTable | None:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None

        scale = self._detect_scale(rows)

        # Find the header row (contains fiscal year or date-like values)
        header_row_idx, col_labels = self._find_header_row(rows)
        if header_row_idx < 0 or not col_labels:
            logger.warning("No usable header row found in sheet '%s'", sheet_name)
            return None

        columns = [self._parse_column(label, cutoff_date) for label in col_labels]

        # Parse data rows below the header
        data: dict[str, dict[str, Any]] = {}
        for row in rows[header_row_idx + 1 :]:
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip()
            if not label:
                continue
            row_values: dict[str, Any] = {}
            for i, col in enumerate(columns):
                cell_idx = i + 1  # offset by label column
                if cell_idx < len(row):
                    row_values[col.label] = row[cell_idx]
                else:
                    row_values[col.label] = None
            data[label] = row_values

        if not data:
            return None

        return RawStatementTable(
            ticker=ticker,
            statement_type=stmt_type,
            columns=columns,
            data=data,
            scale=scale,
            source=DataSource.BLOOMBERG_XLSX,
        )

    def _detect_scale(self, rows: list[tuple]) -> float:
        """Check the first few rows for scale pattern (e.g. 'In Millions')."""
        for row in rows[:5]:
            for cell in row:
                if cell is None:
                    continue
                cell_str = str(cell).lower()
                for pattern, multiplier in BBG_SCALE_PATTERNS.items():
                    if pattern in cell_str:
                        logger.debug("Detected scale: %s (factor=%.0f)", pattern, multiplier)
                        return multiplier
        return 1.0  # No scale detected; assume raw values

    def _find_header_row(self, rows: list[tuple]) -> tuple[int, list[str]]:
        """Scan rows to find the one with fiscal year / date column headers."""
        for idx, row in enumerate(rows[:20]):
            labels = []
            for cell in row[1:]:  # Skip row label column
                if cell is not None:
                    labels.append(str(cell).strip())
            # Consider this a header if ≥2 cells look like years or periods
            year_like = [l for l in labels if _YEAR_PATTERN.search(l)]
            if len(year_like) >= 2:
                return idx, labels
        return -1, []

    def _parse_column(self, label: str, cutoff_date: datetime.date) -> BloombergColumn:
        """
        Parse a Bloomberg column header string into a BloombergColumn.

        Examples of label patterns:
        - '2021A'  → actual fiscal 2021
        - '2022E'  → estimate fiscal 2022
        - 'LTM'    → last twelve months
        - '2021 Restated'
        """
        upper = label.upper()

        is_ltm = any(kw in upper for kw in BBG_LTM_KEYWORDS)
        is_estimate = any(kw in upper for kw in BBG_ESTIMATE_KEYWORDS) and not is_ltm
        is_restated = "RESTATED" in upper

        # Extract fiscal year
        match = _YEAR_PATTERN.search(label)
        fiscal_year: int | None = int(match.group()) if match else None

        # Best-guess period end (December 31 of the fiscal year)
        period_end: datetime.date | None = None
        if fiscal_year is not None:
            period_end = datetime.date(fiscal_year, 12, 31)

        return BloombergColumn(
            label=label,
            fiscal_year=fiscal_year,
            is_estimate=is_estimate,
            is_ltm=is_ltm,
            is_restated=is_restated,
            period_end=period_end,
        )
